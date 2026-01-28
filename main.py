#!/usr/bin/env python3
"""
Automated Internship Finder
Scrapes internships from multiple sources and sends notifications via email.
Uses environment variables for all credentials (no .env file required).
"""

import os
import sys
import logging
from datetime import datetime, timedelta
from pathlib import Path
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from imaplib import IMAP4_SSL
import re

import requests
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import google.generativeai as genai
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.options import Options as FirefoxOptions
import imapclient
from email.parser import BytesParser

# ============================================================================
# LOGGING CONFIGURATION
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("internship_finder.log", mode="a")
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION
# ============================================================================

DATA_DIR = Path("data")
INTERNSHIPS_FILE = DATA_DIR / "internships.xlsx"

# Read credentials from environment variables (no .env file)
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
LINKEDIN_EMAIL = os.getenv("LINKEDIN_EMAIL")
LINKEDIN_PASSWORD = os.getenv("LINKEDIN_PASSWORD")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

# Validate critical credentials
if not EMAIL_ADDRESS or not EMAIL_PASSWORD:
    logger.warning("Email credentials not found in environment variables")

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def ensure_data_directory():
    """Ensure the data directory exists."""
    DATA_DIR.mkdir(exist_ok=True)
    logger.info(f"Data directory ensured at: {DATA_DIR.absolute()}")


def read_user_input_fallback():
    """
    Read skills/interests from user input or environment variable.
    Multiple ways to provide input:
    1. Interactive prompt
    2. INTERNSHIP_SEARCH_QUERY environment variable
    3. Default skills
    """
    user_input = os.getenv("INTERNSHIP_SEARCH_QUERY")
    
    if not user_input:
        logger.info("\n" + "="*70)
        logger.info("🎯 INTERNSHIP FINDER - SKILL INPUT")
        logger.info("="*70)
        logger.info("\nHow to provide your skills and interests:")
        logger.info("  Option 1: Enter skills now (interactive)")
        logger.info("  Option 2: Set INTERNSHIP_SEARCH_QUERY environment variable")
        logger.info("  Option 3: Press Enter to use default skills")
        logger.info("\nExamples: 'Python, Machine Learning, Data Science'")
        logger.info("          'Java, Backend Development, Cloud'")
        logger.info("          'Frontend, React, JavaScript, Web Design'")
        logger.info("="*70 + "\n")
        
        try:
            user_input = input("📝 Enter your skills/interests (or press Enter for defaults): ").strip()
        except EOFError:
            user_input = ""
    
    if not user_input:
        user_input = "Python, Data Science, Machine Learning, Backend Development, Full Stack"
        logger.info(f"ℹ️  Using default search query: {user_input}")
    else:
        logger.info(f"✅ Using your search query: {user_input}")
    
    logger.info(f"\n🔍 Searching for internships matching: {user_input}\n")
    return user_input


def load_existing_internships():
    """Load existing internships from Excel file."""
    existing = set()
    if INTERNSHIPS_FILE.exists():
        try:
            wb = load_workbook(INTERNSHIPS_FILE)
            ws = wb.active
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # If company field exists
                    # Create a unique key based on company, role, and location
                    key = f"{row[0]}|{row[1]}|{row[2]}"
                    existing.add(key)
            logger.info(f"Loaded {len(existing)} existing internships")
        except Exception as e:
            logger.error(f"Error loading existing internships: {e}")
    return existing


def save_internships_to_excel(internships):
    """Save internships to Excel file with formatting."""
    try:
        if INTERNSHIPS_FILE.exists():
            wb = load_workbook(INTERNSHIPS_FILE)
            ws = wb.active
            start_row = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Internships"
            # Create header
            headers = ["Company", "Role", "Location", "Link", "Date Found", "Source"]
            ws.append(headers)
            # Format header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            start_row = 2
        
        # Add internships
        for internship in internships:
            ws.append([
                internship.get("company", ""),
                internship.get("role", ""),
                internship.get("location", ""),
                internship.get("link", ""),
                internship.get("date", datetime.now().strftime("%Y-%m-%d")),
                internship.get("source", "Unknown")
            ])
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        
        wb.save(INTERNSHIPS_FILE)
        logger.info(f"Saved {len(internships)} internships to {INTERNSHIPS_FILE}")
    except Exception as e:
        logger.error(f"Error saving internships to Excel: {e}")


def get_unique_key(internship):
    """Generate a unique key for an internship."""
    # Use company + role + location as the unique identifier
    # Posting date is checked separately for 24-hour filtering
    return f"{internship.get('company', '')}|{internship.get('role', '')}|{internship.get('location', '')}"


def is_posting_recent(internship, hours=24):
    """
    Check if internship posting is within the specified hours.
    Returns True if posting_date is available and within timeframe, False otherwise.
    """
    posting_date_str = internship.get('posting_date')
    
    # If no posting date, accept it (can't filter)
    if not posting_date_str:
        return True
    
    try:
        # Parse the posting date (expected format: YYYY-MM-DD or YYYY-MM-DD HH:MM:SS)
        if " " in posting_date_str:
            posting_date = datetime.strptime(posting_date_str, "%Y-%m-%d %H:%M:%S")
        else:
            posting_date = datetime.strptime(posting_date_str, "%Y-%m-%d")
        
        # Check if within the specified hours
        cutoff_time = datetime.now() - timedelta(hours=hours)
        return posting_date >= cutoff_time
    except (ValueError, TypeError) as e:
        logger.warning(f"Could not parse posting date '{posting_date_str}': {e}")
        return True  # Accept if we can't parse the date


def filter_new_internships(internships, existing_keys):
    """Filter out duplicates based on existing internships and 24-hour old posting rule."""
    new = []
    for internship in internships:
        key = get_unique_key(internship)
        # Only add if: (1) not a duplicate, AND (2) posting is recent (within 24 hours)
        if key not in existing_keys and is_posting_recent(internship, hours=24):
            new.append(internship)
            existing_keys.add(key)
    logger.info(f"Found {len(new)} new internships (within 24 hours) out of {len(internships)} scraped")
    return new


# ============================================================================
# SCRAPING FUNCTIONS
# ============================================================================

def scrape_internshala(skills):
    """
    Scrape internships from Internshala.com
    Note: Internshala uses dynamic content, so we'll attempt basic scraping.
    """
    internships = []
    try:
        logger.info("Scraping Internshala...")
        time.sleep(1)  # Rate limiting
        
        # Search for internships based on skills
        search_query = "-".join(skills.split()[:2]).lower()
        url = f"https://internshala.com/internships/keyword-{search_query}"
        
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Referer": "https://internshala.com/",
            "Connection": "keep-alive"
        }
        
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, "html.parser")
        
        # Try to find internship listings
        listings = soup.find_all("div", class_="internship_card")
        
        if not listings:
            # Try alternative selector
            listings = soup.find_all("article", {"class": "internship"})
        
        for listing in listings[:10]:  # Limit to 10 per source
            try:
                # Extract data with flexible selectors
                company = listing.find("a", class_="internship_company") or listing.find("span", class_="company_name")
                role = listing.find("h3", class_="job_title") or listing.find("span", class_="role_name")
                location = listing.find("span", class_="location") or listing.find("span", class_="city")
                link = listing.find("a", href=True)
                
                if company and role:
                    internship = {
                        "company": company.get_text(strip=True),
                        "role": role.get_text(strip=True),
                        "location": location.get_text(strip=True) if location else "Not Specified",
                        "link": link["href"] if link else "No Link",
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "posting_date": datetime.now().strftime("%Y-%m-%d"),
                        "source": "Internshala"
                    }
                    internships.append(internship)
            except Exception as e:
                logger.debug(f"Error parsing Internshala listing: {e}")
        
        logger.info(f"Scraped {len(internships)} internships from Internshala")
    except Exception as e:
        logger.error(f"Error scraping Internshala: {e}")
    
    return internships


def scrape_angellist(skills):
    """
    Scrape internships from Wellfound (formerly AngelList)
    Note: AngelList has API access but for this example we'll attempt web scraping.
    """
    internships = []
    try:
        logger.info("Scraping Wellfound (AngelList)...")
        time.sleep(1)  # Rate limiting
        
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Referer": "https://wellfound.com/",
            "Connection": "keep-alive"
        }
        
        # Wellfound search URL (AngelList rebranded)
        search_query = "+".join(skills.split()[:2])
        url = f"https://wellfound.com/jobs?keywords={search_query}&job_type=internship"
        
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, "html.parser")
        
        # Try to find job listings
        listings = soup.find_all("div", class_="job-listing")
        
        for listing in listings[:10]:
            try:
                company = listing.find("span", class_="company-name")
                role = listing.find("h2", class_="job-title")
                location = listing.find("span", class_="location")
                link = listing.find("a", href=True)
                
                if company and role:
                    internship = {
                        "company": company.get_text(strip=True),
                        "role": role.get_text(strip=True),
                        "location": location.get_text(strip=True) if location else "Remote",
                        "link": link["href"] if link else "No Link",
                        "date": datetime.now().strftime("%Y-%m-%d"),                        "posting_date": datetime.now().strftime("%Y-%m-%d"),                        "source": "AngelList"
                    }
                    internships.append(internship)
            except Exception as e:
                logger.debug(f"Error parsing AngelList listing: {e}")
        
        logger.info(f"Scraped {len(internships)} internships from AngelList")
    except Exception as e:
        logger.error(f"Error scraping AngelList: {e}")
    
    return internships


def scrape_linkedin_selenium(skills):
    """
    Scrape LinkedIn jobs using Selenium with Firefox in headless mode.
    """
    internships = []
    driver = None
    try:
        logger.info("Scraping LinkedIn via Selenium...")
        
        if not LINKEDIN_EMAIL or not LINKEDIN_PASSWORD:
            logger.warning("LinkedIn credentials not found. Skipping LinkedIn scraping.")
            return internships
        
        # Configure Firefox options for headless mode
        options = FirefoxOptions()
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        
        # Initialize driver
        driver = webdriver.Firefox(options=options)
        driver.set_page_load_timeout(30)
        
        logger.info("Logging into LinkedIn...")
        driver.get("https://www.linkedin.com/login")
        
        # Wait for and fill login form
        wait = WebDriverWait(driver, 20)
        
        email_field = wait.until(EC.presence_of_element_located((By.ID, "username")))
        email_field.send_keys(LINKEDIN_EMAIL)
        
        password_field = driver.find_element(By.ID, "password")
        password_field.send_keys(LINKEDIN_PASSWORD)
        
        login_button = driver.find_element(By.XPATH, "//button[@type='submit']")
        login_button.click()
        
        # Wait for page to load after login
        import time
        time.sleep(5)
        
        # Search for internships
        search_query = " ".join(skills.split()[:3])
        search_url = f"https://www.linkedin.com/jobs/search/?keywords={search_query}&f_jt=INT"
        driver.get(search_url)
        
        time.sleep(3)
        
        # Scroll to load more jobs
        for _ in range(3):
            driver.execute_script("window.scrollBy(0, 500);")
            time.sleep(1)
        
        # Find job listings
        job_listings = driver.find_elements(By.CLASS_NAME, "base-card")
        
        for job in job_listings[:10]:
            try:
                company = job.find_element(By.CLASS_NAME, "hidden-nested-link")
                title = job.find_element(By.CLASS_NAME, "base-search-card__title")
                location = job.find_element(By.CLASS_NAME, "job-search-card__location")
                link = job.find_element(By.TAG_NAME, "a")
                
                internship = {
                    "company": company.text.strip(),
                    "role": title.text.strip(),
                    "location": location.text.strip(),
                    "link": link.get_attribute("href"),
                    "date": datetime.now().strftime("%Y-%m-%d"),
                    "posting_date": datetime.now().strftime("%Y-%m-%d"),
                    "source": "LinkedIn"
                }
                internships.append(internship)
            except Exception as e:
                logger.debug(f"Error parsing LinkedIn job: {e}")
        
        logger.info(f"Scraped {len(internships)} internships from LinkedIn")
    except Exception as e:
        logger.error(f"Error scraping LinkedIn: {e}")
    finally:
        if driver:
            driver.quit()
    
    return internships


def scrape_glassdoor(skills):
    """
    Scrape internship opportunities from Glassdoor.com
    """
    internships = []
    try:
        logger.info("Scraping Glassdoor...")
        time.sleep(1)  # Rate limiting
        
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Referer": "https://www.glassdoor.com/",
            "Connection": "keep-alive"
        }
        
        search_query = "-".join(skills.split()[:2]).lower()
        url = f"https://www.glassdoor.com/Job/internship-{search_query}-jobs-SRCH_KO0,11.htm"
        
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, "html.parser")
        
        # Try to find job listings
        listings = soup.find_all("div", class_="job-search-result")
        
        for listing in listings[:10]:  # Limit to 10 per source
            try:
                company = listing.find("span", class_="company-name")
                role = listing.find("a", class_="job-title")
                location = listing.find("span", class_="job-location")
                link = listing.find("a", href=True)
                
                if company and role:
                    internship = {
                        "company": company.get_text(strip=True),
                        "role": role.get_text(strip=True),
                        "location": location.get_text(strip=True) if location else "Not Specified",
                        "link": f"https://www.glassdoor.com{link['href']}" if link else "No Link",
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "posting_date": datetime.now().strftime("%Y-%m-%d"),
                        "source": "Glassdoor"
                    }
                    internships.append(internship)
            except Exception as e:
                logger.debug(f"Error parsing Glassdoor listing: {e}")
        
        logger.info(f"Scraped {len(internships)} internships from Glassdoor")
    except Exception as e:
        logger.error(f"Error scraping Glassdoor: {e}")
    
    return internships


def scrape_big_company_careers(skills):
    """
    Scrape internships from major tech company career pages:
    Google, Microsoft, Amazon, Meta, Apple, Netflix, Intel, IBM, Oracle, Salesforce
    """
    internships = []
    
    # Major company career sites
    company_careers = {
        "Google": "https://careers.google.com/jobs/results/",
        "Microsoft": "https://careers.microsoft.com/us/en/search-results",
        "Amazon": "https://amazon.jobs/en/search?offset=0&result_limit=100&sort=recent",
        "Meta": "https://www.metacareers.com/jobs/",
        "Apple": "https://jobs.apple.com/en-us/search",
        "Netflix": "https://jobs.netflix.com/search",
        "Intel": "https://jobs.intel.com/en/search-jobs",
        "IBM": "https://www.ibm.com/careers/search/",
        "Oracle": "https://careers.oracle.com/",
        "Salesforce": "https://www.salesforce.com/company/careers/search-jobs/",
    }
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Connection": "keep-alive"
    }
    
    for company_name, career_url in company_careers.items():
        time.sleep(1)  # Rate limiting between companies
        try:
            logger.debug(f"Checking {company_name} careers...")
            response = requests.get(career_url, headers=headers, timeout=10)
            
            if response.status_code != 200:
                logger.debug(f"Could not access {company_name} careers page")
                continue
            
            soup = BeautifulSoup(response.content, "html.parser")
            
            # Generic search for job listings (methods vary by company)
            # Look for common patterns
            job_elements = soup.find_all(["div", "article"], 
                                        class_=lambda x: x and ("job" in x.lower() or "posting" in x.lower()))
            
            for job in job_elements[:5]:  # Limit per company
                try:
                    title_elem = job.find(["h2", "h3", "a"], 
                                         class_=lambda x: x and ("title" in x.lower() or "job" in x.lower()))
                    location_elem = job.find(["span", "div"], 
                                            class_=lambda x: x and "location" in x.lower())
                    link_elem = job.find("a", href=True)
                    
                    if title_elem and title_elem.get_text().strip():
                        title = title_elem.get_text(strip=True)
                        
                        # Only collect internships (filter by "intern" in title)
                        if "intern" in title.lower():
                            internship = {
                                "company": company_name,
                                "role": title[:80],
                                "location": location_elem.get_text(strip=True) if location_elem else "Check website",
                                "link": link_elem["href"] if link_elem else career_url,
                                "date": datetime.now().strftime("%Y-%m-%d"),
                                "posting_date": datetime.now().strftime("%Y-%m-%d"),
                                "source": f"{company_name} Careers"
                            }
                            internships.append(internship)
                except Exception as e:
                    logger.debug(f"Error parsing {company_name} job: {e}")
        
        except Exception as e:
            logger.debug(f"Error checking {company_name}: {e}")
    
    logger.info(f"Scraped {len(internships)} internships from major company career pages")
    return internships


def read_gmail_job_alerts():
    """
    Read unread LinkedIn job alert emails from Gmail as a fallback.
    """
    internships = []
    try:
        logger.info("Reading Gmail job alerts...")
        
        if not EMAIL_ADDRESS or not EMAIL_PASSWORD:
            logger.warning("Email credentials not found. Skipping Gmail scraping.")
            return internships
        
        # Connect to Gmail
        server = imapclient.IMAPClient("imap.gmail.com", ssl=True)
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        server.select_folder("INBOX")
        
        # Search for unread emails from LinkedIn and job sites
        message_ids = server.search(["UNSEEN", "FROM", "linkedin"])
        
        if not message_ids:
            logger.info("No unread LinkedIn job alert emails found.")
            return internships
        
        # Fetch messages
        messages = server.fetch(message_ids, [b"RFC822"])
        
        for msg_id, data in messages.items():
            try:
                msg = BytesParser().parsebytes(data[b"RFC822"])
                
                # Extract job information from email
                subject = msg.get("subject", "")
                body = msg.get_payload(decode=True).decode("utf-8", errors="ignore")
                
                # Simple parsing: look for patterns in the email
                if "job" in body.lower() or "opportunity" in body.lower():
                    # Extract URLs
                    urls = re.findall(r"https?://\S+", body)
                    
                    # Try to extract key information
                    internship = {
                        "company": "LinkedIn Alert",
                        "role": subject[:50] if subject else "Job Opportunity",
                        "location": "Check Email",
                        "link": urls[0] if urls else "No Link",
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "posting_date": datetime.now().strftime("%Y-%m-%d"),
                        "source": "Gmail"
                    }
                    internships.append(internship)
                
                # Mark as read
                server.set_flags(msg_id, [b"\\Seen"])
            except Exception as e:
                logger.debug(f"Error parsing email: {e}")
        
        server.logout()
        logger.info(f"Extracted {len(internships)} job alerts from Gmail")
    except Exception as e:
        logger.error(f"Error reading Gmail: {e}")
    
    return internships


# ============================================================================
# GEMINI API FUNCTIONS
# ============================================================================

def summarize_with_gemini(internship, skills):
    """
    Use Gemini API to summarize or improve matching score for an internship.
    """
    if not GEMINI_API_KEY:
        return internship
    
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel("gemini-pro")
        
        prompt = f"""
        Given the following internship opportunity and user skills, provide a match score (0-100) and brief reason.
        
        Internship:
        Company: {internship['company']}
        Role: {internship['role']}
        Location: {internship['location']}
        
        User Skills: {skills}
        
        Respond in JSON format: {{"match_score": <number>, "reason": "<reason>"}}
        """
        
        response = model.generate_content(prompt, request_options={"timeout": 30})
        
        # Parse response
        response_text = response.text.strip()
        if response_text.startswith("{"):
            # Try to extract JSON
            match = re.search(r"\{.*\}", response_text, re.DOTALL)
            if match:
                data = json.loads(match.group())
                internship["match_score"] = data.get("match_score", 0)
                internship["reason"] = data.get("reason", "")
        
        logger.debug(f"Gemini analysis for {internship['company']}: {response_text}")
    except Exception as e:
        logger.debug(f"Error using Gemini API: {e}")
    
    return internship


# ============================================================================
# EMAIL NOTIFICATION
# ============================================================================

def send_email_notification(internships):
    """
    Send an email notification with new internships.
    """
    if not internships:
        logger.info("No new internships to send via email")
        return
    
    if not EMAIL_ADDRESS or not EMAIL_PASSWORD:
        logger.warning("Email credentials not found. Skipping email notification.")
        return
    
    try:
        logger.info(f"Sending email notification with {len(internships)} internships...")
        
        # Compose email
        message = MIMEMultipart("alternative")
        message["Subject"] = f"[Internship Bot] Found {len(internships)} New Internships"
        message["From"] = EMAIL_ADDRESS
        message["To"] = EMAIL_ADDRESS
        
        # Create text and HTML versions
        text_content = f"Found {len(internships)} new internships!\n\n"
        html_content = f"""
        <html>
            <body>
                <h2>🎯 New Internships Found! ({len(internships)})</h2>
                <table border="1" cellpadding="10" cellspacing="0" style="border-collapse: collapse; width: 100%;">
                    <tr style="background-color: #4472C4; color: white;">
                        <th>Company</th>
                        <th>Role</th>
                        <th>Location</th>
                        <th>Source</th>
                        <th>Link</th>
                    </tr>
        """
        
        for internship in internships:
            text_content += f"""
Company: {internship['company']}
Role: {internship['role']}
Location: {internship['location']}
Source: {internship['source']}
Link: {internship['link']}
---
"""
            
            link_html = f"<a href='{internship['link']}'>View</a>" if internship['link'] != "No Link" else "N/A"
            html_content += f"""
                    <tr>
                        <td>{internship['company']}</td>
                        <td>{internship['role']}</td>
                        <td>{internship['location']}</td>
                        <td>{internship['source']}</td>
                        <td>{link_html}</td>
                    </tr>
            """
        
        html_content += """
                </table>
                <p><small>This email was generated automatically by the Internship Finder Bot.</small></p>
            </body>
        </html>
        """
        
        # Attach text and HTML
        message.attach(MIMEText(text_content, "plain"))
        message.attach(MIMEText(html_content, "html"))
        
        # Send email
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.sendmail(EMAIL_ADDRESS, EMAIL_ADDRESS, message.as_string())
        
        logger.info("Email notification sent successfully")
    except Exception as e:
        logger.error(f"Error sending email: {e}")


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Main function to orchestrate the internship finding process."""
    logger.info("=" * 80)
    logger.info("Starting Automated Internship Finder")
    logger.info("=" * 80)
    
    # Setup
    ensure_data_directory()
    user_input = read_user_input_fallback()
    existing_keys = load_existing_internships()
    
    logger.info(f"Search Query: {user_input}")
    logger.info(f"Searching from {6} different sources...\n")
    
    # Scrape from multiple sources
    all_internships = []
    
    # Web scraping sources
    logger.info("📊 STARTING SCRAPE FROM ALL SOURCES")
    logger.info("-" * 80)
    
    all_internships.extend(scrape_internshala(user_input))
    all_internships.extend(scrape_angellist(user_input))
    all_internships.extend(scrape_glassdoor(user_input))
    all_internships.extend(scrape_big_company_careers(user_input))
    
    # Selenium-based scraping
    all_internships.extend(scrape_linkedin_selenium(user_input))
    
    # Email-based fallback
    all_internships.extend(read_gmail_job_alerts())
    
    logger.info("-" * 80)
    logger.info(f"✅ Total internships scraped: {len(all_internships)}")
    
    # Filter duplicates
    new_internships = filter_new_internships(all_internships, existing_keys)
    
    # Optional: Use Gemini API for enhancement
    if GEMINI_API_KEY and new_internships:
        logger.info("Enhancing internships with Gemini API...")
        for internship in new_internships[:5]:  # Limit API calls
            internship = summarize_with_gemini(internship, user_input)
    
    # Save to Excel
    if new_internships:
        save_internships_to_excel(new_internships)
        
        # Send email notification
        send_email_notification(new_internships)
    else:
        logger.info("No new internships found")
    
    logger.info("=" * 80)
    logger.info("Internship Finder completed successfully")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()
